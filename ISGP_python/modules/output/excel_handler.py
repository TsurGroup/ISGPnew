import os
import numpy as np
import openpyxl
import pandas as pd


from config.version import VERSION
from file_managment.file_manager import get_excel_path
from cache.cache import get_experiment_data, get_project_constants_from_cache
from models.genome import Genome

def save_run_to_excel(run_num, model: Genome):

    project_constants = get_project_constants_from_cache()
    solution_samples = model.get_genome_value(project_constants.time_samples)
    experiment_data = get_experiment_data(0)

    point = -(experiment_data.logarithmic_relaxation_time[0] - experiment_data.logarithmic_relaxation_time[1]) / 3
    normalization_factor = max(experiment_data.real_impedance)
    print(normalization_factor)
    area = model.get_area(project_constants.time_samples, point)

    # Prepare the data for the current run using the 'model' argument
    data_run = {
        'Frequency': experiment_data.frequency, 
        'Z_Real': experiment_data.real_impedance,
        'Z_Imaginary': experiment_data.imaginary_impedance
    }
    
    # Create a DataFrame for experiment data
    df_run = pd.DataFrame(data_run)
    df_run[''] = ''  # Empty unnamed column

    # Create a DataFrame for model data
    model_data = {
        'log(tau)': project_constants.time_samples, 
        'DFRT': solution_samples,
    }
    df_model = pd.DataFrame(model_data)  # Convert model data to a DataFrame
    df_model[''] = ''  # Empty unnamed column

    # Create DataFrame for model properties
    model_properties = {
        'Properties': ['Compatibility:', 'Discrepancy:', 'Area:'],
        ' ': [model.new_fitness, model.discrepancy, area]  # Empty unnamed column
    }
    df_model_properties = pd.DataFrame(model_properties)
    df_model_properties[''] = ''  # Empty unnamed column

    # Prepare peak data
    peak_data = {
        'Peak': [],
        'Peak Type': [],
        'Area': [],
        'Effective R': [],
        'log(Tau)': [],
        'Tau': [],
        'Frequency': [],
        'Effective C': [],
    }
    
    for i in range(len(model.functions)):
        peak_data['Peak'].append(f'Peak {i+1}')
        peak_data['Peak Type'].append(model.functions[i].func_type.description)
        peak_area = model.functions[i].get_area(project_constants.time_samples, point)
        peak_data['Area'].append(peak_area)
        peak_data['Effective R'].append(peak_area * normalization_factor)
        log_tau = model.functions[i].mean
        peak_data['log(Tau)'].append(log_tau)# not best practice
        peak_data['Tau'].append(10 ** log_tau)
        peak_data['Frequency'].append(1 / (2 * np.pi * (10 ** log_tau)))
        peak_data['Effective C'].append(10 ** log_tau/(peak_area * normalization_factor))

    df_peak_data = pd.DataFrame(peak_data)

    # Concatenate experimental data, model data, and model properties along the columns (side by side)
    df_combined = pd.concat([df_run, df_model, df_model_properties, df_peak_data], axis=1)

    excel_path = get_excel_path()

    # If the Excel file exists, load it; if not, create a new workbook
    if not os.path.exists(excel_path):
        # First run: create a new workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f"Run {run_num + 1}"  # Start from "num 1"
        workbook.save(excel_path)
    else:
        # If the workbook exists, load it
        workbook = openpyxl.load_workbook(excel_path)
    
    # Append new sheet
    new_sheet_name = f"Run {run_num + 1}"
    if new_sheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=new_sheet_name)
    
    # Save the workbook to retain the new sheet
    workbook.save(excel_path)

    # Use pandas ExcelWriter to append data to the correct sheet
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        # Write the combined DataFrame to the new sheet without headers
        df_combined.to_excel(writer, sheet_name=new_sheet_name, index=False, header=True)

    print(f"Data saved to sheet '{new_sheet_name}' in '{excel_path}'.")

    return 1


def save_metadata():
    """
    Save the program version to the metadata sheet in the Excel file.
    Creates the metadata sheet if it doesn't exist and ensures it appears last.
    """
    metadata_sheet_name = "Metadata"

    # Check if the Excel file exists
    excel_path = get_excel_path()
    workbook = openpyxl.load_workbook(excel_path)
    
    metadata_sheet = workbook.create_sheet(title=metadata_sheet_name)
    # Prepare metadata data
    metadata_data = {"Field": ["Program Version"], "Value": [VERSION]}
    df_metadata = pd.DataFrame(metadata_data)

    # Write metadata data to the sheet
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_metadata.to_excel(writer, sheet_name=metadata_sheet_name, index=False)

    print(f"Program version '{VERSION}' saved to the metadata sheet.")
